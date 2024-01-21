import requests
from bs4 import BeautifulSoup
import openpyxl
import json
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from collections import Counter
import re

url = 'https://insights.blackcoffer.com/rising-it-cities-and-its-impact-on-the-economy-environment-infrastructure-and-city-life-by-the-year-2040-2/'

response = requests.get(url)
html_content = response.content

soup = BeautifulSoup(html_content, "html.parser")


#the XLSX file
wb = openpyxl.load_workbook("Input.xlsx")
sheet = wb["Sheet1"]

# Extracting links from the sheet
article_links = []
for row in sheet.iter_rows(min_row=2):  # Skip header row
    # Access the second column for links
    article_link = row[1].value
    article_links.append(article_link)

def extract_article_data(article_url):
  # Fetch the article content
  article_response = requests.get(article_url)
  article_content = article_response.content

  # Create BeautifulSoup object for the article
  article_soup = BeautifulSoup(article_content, "html.parser")

  # Extract title
  title_element = article_soup.find("h1", class_=["entry-title", "tdb-title-text"])
  if title_element:
    title = title_element.text.strip()
  else:
    title = "" # Handle missing title gracefully

  # Extract text
  text_parts = []
  for p in article_soup.find_all("p", class_=None):  # Select p tags with no class
    text_parts.append(p.text.strip())

  # Join all extracted text parts
  text = "".join(text_parts)

  # Return extracted data
  return {"title": title, "text": text}

c = 0
articles_data = []
for article_link in article_links:
  try:
    article_data = extract_article_data(article_link)
    articles_data.append(article_data)
  except Exception as e:
    print(f"Error extracting data from article: {article_link}")
    print(f"Error message: {e}")
  print('Data Extracted')


with open("extracted_articles.json", "w") as f:
  json.dump(articles_data, f, indent=4)

print('Done Extracting.')


# NLP task

wb = openpyxl.load_workbook(r'C:\Users\Jayant Shekhawat\Output Data Structure.xlsx')
sheet = wb.active

#Columns
word_count_col = 12
average_words_per_sentence_col = 10
complex_word_count_col = 11
personal_pronoun_count_col = 5
syllables_per_word_col = 13
positive_score_col = 3
negative_score_col = 4
polarity_score_col = 5
subjectivity_score_col = 6
fog_index_col = 9
average_sentence_length_col = 7
percentage_complex_words_col = 8
average_word_length_col = 15

current_row = 2



with open("extracted_articles.json", "r") as f:
  article_data = json.load(f)

stop_words = set()
stopword_file_paths = [r"C:\Users\Jayant Shekhawat\StopWords_Auditor.txt", r"C:\Users\Jayant Shekhawat\StopWords_Currencies.txt", r"C:\Users\Jayant Shekhawat\StopWords_Generic.txt", r"C:\Users\Jayant Shekhawat\StopWords_GenericLong.txt", r"C:\Users\Jayant Shekhawat\StopWords_DatesandNumbers.txt", r"C:\Users\Jayant Shekhawat\StopWords_Geographic.txt", r"C:\Users\Jayant Shekhawat\StopWords_Names.txt"]
for file_path in stopword_file_paths:
  with open(file_path, "r") as f:
    for line in f:
      stop_words.add(line.strip())


positive_words = set()
negative_words = set()
# Read positive words
with open(r"C:\Users\Jayant Shekhawat\positive-words.txt", "r") as f:
  for word in f.readlines():  # Read and add each line to the set
    positive_words.add(word.strip())
# Read negative words
with open(r"C:\Users\Jayant Shekhawat\negative-words.txt", "r") as f:
  for word in f.readlines():
    negative_words.add(word.strip())


def count_syllables(word):
  """
  Counts the number of syllables in a word.

  Args:
    word: The word to count syllables for.

  Returns:
    The number of syllables in the word.
  """
  # Define vowel sounds
  vowels = "aeiouy"

  # Count syllables
  syllables = 0
  for i in range(len(word)):
    if word[i] in vowels:
      if i == 0 or word[i-1] not in vowels:
        syllables += 1

  # Handle exceptions
  if word.endswith("es") or word.endswith("ed"):
    syllables -= 1

  return syllables

for article in article_data:
  # Initialize variables
  positive_score = 0
  negative_score = 0
  total_words = 0
  total_words_cleaned = 0
  complex_words = 0
  sentences = 0
  fog_index = 0.0
  average_words_per_sentence = 0.0
  personal_pronoun_count = 0
  average_word_length = 0
  total_characters = 0
  total_sentences_cleaned = 0
  syllables_per_word = []
  
  clean_text = ""
  for word in word_tokenize(article["text"]):
    if word not in stop_words:
      clean_text += word + " "
  
  # Calculate sentiment scores
  for word in word_tokenize(clean_text):
    total_words_cleaned += 1
    if word in positive_words:
      positive_score += 1
    if word in negative_words:
      negative_score += 1

  # Calculate polarity and subjectivity scores
  polarity_score = (positive_score - negative_score) / ((positive_score + negative_score) + 0.000001)
  subjectivity_score = (positive_score + negative_score) / (total_words_cleaned + 0.000001)

  # Calculating total words
  for word in word_tokenize(article['text']):
    total_words += 1
  
  # Calculate average word length
  for word in word_tokenize(article["text"].lower()):
    total_characters += len(word)
  average_word_length = total_characters/total_words if total_words > 0 else 0

  # Define the regex pattern for personal pronouns
  personal_pronoun_pattern = r"\b(I|we|We|My|my|Ours|Our|our|ours|us|Us)\b(?! US)"

  # Count personal pronouns
  for word in word_tokenize(article["text"]):
      # Check if the word matches the personal pronoun pattern
      if re.match(personal_pronoun_pattern, word):
        personal_pronoun_count += 1

  # Count syllables per word
  for word in word_tokenize(article['text']):
    syllables_per_word.append(count_syllables(word))

  # Count Complex Words
  for word in word_tokenize(clean_text):
    syl_count = count_syllables(word)
    if syl_count > 2:
      complex_words += 1

  # Calculate number of sentences
  sentences = len(nltk.sent_tokenize(article["text"]))

  # Calculate the number of sentences in cleaned text
  total_sentences_cleaned = len(nltk.sent_tokenize(clean_text))

  # Calculate average sentence length
  average_words_per_sentence = total_words / sentences if sentences > 0 else 0

  # Calculate other metrics
  average_sentence_length = total_words_cleaned / total_sentences_cleaned if total_sentences_cleaned > 0 else 0
  percentage_complex_words = (complex_words / total_words_cleaned) * 100 if total_words_cleaned > 0 else 0
  fog_index = 0.4 * (average_sentence_length + percentage_complex_words)

  # Print results
  print("Article Title:", article["title"])
  print("Word Count:", total_words)
  print("Average Words Per Sentence:", average_words_per_sentence)
  print("Average Sentence Length: ", average_sentence_length)
  print('Percentage of Complex Words:', percentage_complex_words)
  print("Complex Word Count:", complex_words)
  print("Personal Pronoun Count:", personal_pronoun_count)
  print("Syllables per Word:", syllables_per_word)
  print("Positive Score:", positive_score)
  print("Negative Score:", negative_score)
  print("Polarity Score:", polarity_score)
  print("Subjectivity Score:", subjectivity_score)
  print("Gunning Fog Index:", fog_index)
  print("---")

  syllables_string = ",".join(map(str, syllables_per_word))
  sheet.cell(row=current_row, column=syllables_per_word_col).value = syllables_string
  sheet.cell(row=current_row, column=word_count_col).value = total_words
  sheet.cell(row=current_row, column=average_words_per_sentence_col).value = average_words_per_sentence
  sheet.cell(row=current_row, column=complex_word_count_col).value = complex_words
  sheet.cell(row=current_row, column=personal_pronoun_count_col).value = personal_pronoun_count
  sheet.cell(row=current_row, column=positive_score_col).value = positive_score
  sheet.cell(row=current_row, column=negative_score_col).value = negative_score
  sheet.cell(row=current_row, column=polarity_score_col).value = polarity_score
  sheet.cell(row=current_row, column=subjectivity_score_col).value = subjectivity_score
  sheet.cell(row=current_row, column=average_sentence_length_col).value = average_sentence_length
  sheet.cell(row=current_row, column=percentage_complex_words_col).value = percentage_complex_words
  sheet.cell(row=current_row, column=fog_index_col).value = fog_index
  sheet.cell(row=current_row, column=average_word_length_col).value = average_word_length

  current_row += 1

wb.save('Output Data Structure.xlsx')


